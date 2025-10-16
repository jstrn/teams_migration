#!/usr/bin/env python3
"""
classify_and_report.py
Department classification and Excel report generation for SharePoint migration analysis.
"""
# /// script
# dependencies = ["pandas", "openpyxl"]
# ///

import argparse
import json
import csv
import re
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Embedded department keywords for classification
DEPARTMENT_KEYWORDS = {
    "Executive": ["executive", "ceo", "president", "c-suite", "board", "exec", "strategy", "leadership", "founders", "management team"],
    "Finance": ["finance", "financials", "budgets", "forecasting", "expense reports", "ap", "ar", "accounts payable", "accounts receivable", "invoicing", "payments", "cashflow", "general ledger", "reconciliation"],
    "Accounting": ["accounting", "bookkeep", "bookkeeping", "journal entries", "audit", "balance sheet", "profit loss", "trial balance", "tax", "1099", "w2", "payroll journal", "qb", "quickbooks", "myob", "sage", "peachtree"],
    "Human Resources": ["hr", "human resources", "employees", "staff records", "recruiting", "hiring", "onboarding", "offboarding", "benefits", "personnel", "timesheets", "attendance", "performance reviews", "resumes", "applicants", "training records"],
    "Payroll": ["payroll", "paystubs", "payslips", "compensation", "salary", "wages", "deductions", "timecards", "adp", "paychex", "intuit payroll"],
    "Legal": ["legal", "contracts", "agreements", "nda", "litigation", "compliance docs", "policies", "terms", "licensing", "trademarks", "patents", "regulatory", "counsel"],
    "Administration": ["admin", "administration", "office admin", "clerical", "schedules", "forms", "procedures", "templates", "policies", "correspondence", "memos"],
    "Operations": ["operations", "ops", "procedures", "workflow", "process docs", "sop", "production schedule", "maintenance", "scheduling", "dispatch", "capacity planning"],
    "Facilities": ["facilities", "maintenance", "building", "property", "lease", "office layout", "safety", "security", "janitorial", "equipment logs", "hvac", "inspection"],
    "Information Technology": ["it", "information technology", "network", "systems", "infrastructure", "servers", "backups", "helpdesk", "support tickets", "software", "hardware", "configs", "logs", "endpoints", "security", "vpn", "firewall", "group policy"],
    "Security": ["security", "infosec", "cyber", "antivirus", "monitoring", "soc", "incidents", "breaches", "policies", "credentials", "encryption", "risk assessment"],
    "Sales": ["sales", "crm", "leads", "opportunities", "quotes", "proposals", "orders", "customers", "pipelines", "forecasts", "commissions", "sf", "salesforce"],
    "Marketing": ["marketing", "campaigns", "ads", "advertising", "branding", "collateral", "flyers", "social media", "website", "seo", "sem", "newsletters", "content", "creative", "graphics", "media", "events"],
    "Customer Service": ["customer service", "support", "tickets", "cases", "complaints", "feedback", "returns", "helpdesk", "customers", "satisfaction", "warranty"],
    "Product Management": ["product", "roadmap", "features", "specs", "backlog", "release notes", "planning", "requirements", "prd", "epic", "jira exports"],
    "Project Management": ["projects", "pm", "gantt", "schedule", "timelines", "milestones", "deliverables", "scope", "pmo", "client projects", "resources"],
    "Procurement": ["procurement", "purchasing", "orders", "suppliers", "vendors", "requisitions", "po", "purchase orders", "quotes", "bids", "sourcing"],
    "Supply Chain": ["supply chain", "logistics", "inventory", "fulfillment", "distribution", "warehouse", "shipping", "receiving", "tracking", "supply", "demand"],
    "Inventory": ["inventory", "stock", "sku", "items", "parts", "assets", "warehouse", "supplies", "bins", "counts", "reorder", "inventory report"],
    "Logistics": ["logistics", "shipping", "delivery", "freight", "transport", "routing", "trucking", "carriers", "dispatch", "packing slips", "manifests"],
    "Engineering": ["engineering", "design", "cad", "drawings", "blueprints", "schematics", "calculations", "specs", "r&d", "prototype", "testing", "simulation"],
    "Research and Development": ["research", "development", "rnd", "lab", "experiments", "innovation", "patents", "trials", "formulations", "testing", "reports"],
    "Quality Assurance": ["qa", "qc", "quality", "inspection", "defects", "audits", "standards", "testing", "nonconformance", "corrective actions", "iso", "reports"],
    "Training": ["training", "learning", "lms", "education", "onboarding", "tutorials", "certification", "course", "development plan", "safety training"],
    "Business Development": ["business dev", "bizdev", "partnerships", "opportunities", "mergers", "alliances", "joint ventures", "proposals", "growth"],
    "Vendor Management": ["vendor", "supplier", "partner", "agreements", "onboarding", "evaluations", "scorecards", "compliance docs", "purchase orders"],
    "Compliance": ["compliance", "policy", "audit", "regulation", "iso", "soc2", "gdpr", "hipaa", "certification", "documentation", "risk register", "standards"]
}


class DepartmentClassifier:
    """Implements bottom-up consensus algorithm for department classification."""
    
    def __init__(self, keywords_file=None):
        """
        Initialize classifier with embedded or external keywords.
        
        Args:
            keywords_file: Optional CSV file to load keywords from. If None, uses embedded keywords.
        """
        if keywords_file and Path(keywords_file).exists():
            self.department_keywords = {}
            self.load_keywords_from_file(keywords_file)
        else:
            # Use embedded keywords
            self.department_keywords = DEPARTMENT_KEYWORDS
            print(f"Using embedded keywords for {len(self.department_keywords)} departments")
    
    def load_keywords_from_file(self, keywords_file):
        """Load and parse department keywords from CSV file."""
        print(f"Loading department keywords from {keywords_file}...")
        
        with open(keywords_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                dept = row['Department'].strip()
                keywords_str = row['Keywords'].strip()
                
                # Parse keywords (comma-separated, may include spaces and underscores)
                keywords = [kw.strip().lower().replace('_', ' ') 
                           for kw in keywords_str.split(',')]
                
                self.department_keywords[dept] = keywords
        
        print(f"Loaded {len(self.department_keywords)} departments from file")
    
    def tokenize(self, text):
        """Convert text to tokens for keyword matching."""
        if not text:
            return []
        
        # Convert to lowercase, replace special chars with spaces
        text = text.lower()
        text = re.sub(r'[_\-\.\\/]+', ' ', text)
        
        # Extract words (alphanumeric sequences)
        tokens = re.findall(r'\b\w+\b', text)
        
        return tokens
    
    def score_text(self, text, department_keywords):
        """Score text against department keywords."""
        tokens = self.tokenize(text)
        
        if not tokens:
            return 0
        
        score = 0
        tokens_str = ' '.join(tokens)
        
        for keyword in department_keywords:
            keyword_tokens = keyword.split()
            
            # Check for exact keyword match in token string
            if keyword in tokens_str:
                score += len(keyword_tokens) * 2  # Multi-word matches get bonus
            else:
                # Check individual keyword tokens
                for kw_token in keyword_tokens:
                    if kw_token in tokens:
                        score += 1
        
        return score
    
    def classify_item(self, name, parent_path=''):
        """Classify a single item (file or folder) by name."""
        scores = {}
        
        for dept, keywords in self.department_keywords.items():
            # Score the item name
            name_score = self.score_text(name, keywords)
            
            # Score the parent path (lower weight)
            path_score = self.score_text(parent_path, keywords) * 0.3
            
            scores[dept] = name_score + path_score
        
        # Return department with highest score
        if max(scores.values()) > 0:
            return max(scores, key=scores.get), scores
        else:
            return "Unclassified", scores
    
    def classify_hierarchy(self, df):
        """
        Implement bottom-up consensus algorithm.
        
        Algorithm:
        1. For each folder from deepest to shallowest:
            a. Score files within folder
            b. Score subfolder classifications
            c. Score folder name itself (higher weight)
            d. Aggregate scores and assign department
        2. Propagate classifications upward
        """
        print("\nStarting hierarchical classification...")
        
        # Prepare data structures
        folders_df = df[df['Type'] == 'Folder'].copy()
        files_df = df[df['Type'] == 'File'].copy()
        
        # Calculate folder depths
        folders_df['Depth'] = folders_df['Path'].apply(lambda x: x.count('\\'))
        
        # Sort by depth (deepest first)
        folders_df = folders_df.sort_values('Depth', ascending=False)
        
        # Dictionary to store folder classifications and scores
        folder_classifications = {}
        folder_confidence = {}
        
        print(f"Classifying {len(folders_df)} folders...")
        
        # Process each folder from deepest to shallowest
        for idx, folder_row in folders_df.iterrows():
            folder_path = folder_row['Path']
            folder_name = folder_row['Name']
            
            # Initialize department scores
            dept_scores = defaultdict(float)
            
            # 1. Score folder name (weight: 5.0)
            folder_dept, folder_scores = self.classify_item(folder_name)
            for dept, score in folder_scores.items():
                dept_scores[dept] += score * 5.0
            
            # 2. Score files in this folder (weight: 1.0 per file)
            folder_files = files_df[files_df['Path'].str.startswith(folder_path + '\\')]
            for _, file_row in folder_files.iterrows():
                file_name = file_row['Name']
                file_dept, file_scores = self.classify_item(file_name)
                for dept, score in file_scores.items():
                    dept_scores[dept] += score * 1.0
            
            # 3. Score immediate subfolders (weight: 3.0 per subfolder)
            immediate_subfolders = [fp for fp in folder_classifications.keys() 
                                   if Path(fp).parent == Path(folder_path)]
            
            for subfolder_path in immediate_subfolders:
                subfolder_dept = folder_classifications[subfolder_path]
                if subfolder_dept != "Unclassified":
                    # Give score to the subfolder's assigned department
                    dept_scores[subfolder_dept] += folder_confidence.get(subfolder_path, 1.0) * 3.0
            
            # Determine final classification
            if dept_scores and max(dept_scores.values()) > 0:
                best_dept = max(dept_scores, key=dept_scores.get)
                confidence = dept_scores[best_dept]
            else:
                best_dept = "Unclassified"
                confidence = 0.0
            
            folder_classifications[folder_path] = best_dept
            folder_confidence[folder_path] = confidence
        
        # Classify files based on their parent folder
        print(f"Classifying {len(files_df)} files based on folder hierarchy...")
        
        file_classifications = {}
        for idx, file_row in files_df.iterrows():
            file_path = file_row['Path']
            parent_folder = str(Path(file_path).parent)
            
            # Find the closest parent folder that has a classification
            current_path = parent_folder
            classification = "Unclassified"
            
            while current_path:
                if current_path in folder_classifications:
                    classification = folder_classifications[current_path]
                    break
                
                parent = str(Path(current_path).parent)
                if parent == current_path:  # Reached root
                    break
                current_path = parent
            
            file_classifications[file_path] = classification
        
        print("Classification complete!")
        
        return folder_classifications, folder_confidence, file_classifications


class ExcelReportGenerator:
    """Generate formatted Excel workbook with multiple sheets."""
    
    def __init__(self, output_file):
        self.output_file = output_file
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Define styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def apply_header_style(self, ws, row=1):
        """Apply header styling to the first row."""
        for cell in ws[row]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def auto_fit_columns(self, ws, min_width=10, max_width=50):
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, df, folder_classifications, file_classifications):
        """Create summary sheet with high-level statistics."""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Calculate statistics
        total_files = len(df[df['Type'] == 'File'])
        total_folders = len(df[df['Type'] == 'Folder'])
        total_size = df[df['Type'] == 'File']['SizeBytes'].sum()
        
        # Convert size to human-readable format
        def format_size(bytes_val):
            for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
                if bytes_val < 1024:
                    return f"{bytes_val:.2f} {unit}"
                bytes_val /= 1024
            return f"{bytes_val:.2f} PB"
        
        # Count issues
        issues = {
            'Long Paths (>255 chars)': df['IsTooLongPath'].sum(),
            'Large Files (>50MB)': df['IsLargeFile'].sum(),
            'Folders with >20k files': df['IsTooManyFiles'].sum(),
            'Unsupported Characters': df['HasUnsupportedChars'].sum(),
            'Unsafe File Extensions': df['IsUnsafeExtension'].sum(),
            'Explicit Permissions': df['HasExplicitPermissions'].sum()
        }
        
        # Count by department
        dept_counts = Counter(file_classifications.values())
        
        # Write summary data
        row = 1
        ws.cell(row, 1, "SharePoint Migration Analysis - Summary Report")
        ws.cell(row, 1).font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row, 1, "Generated:")
        ws.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 2
        
        # Overall statistics
        ws.cell(row, 1, "Overall Statistics")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row, 1, "Total Files:")
        ws.cell(row, 2, total_files)
        row += 1
        
        ws.cell(row, 1, "Total Folders:")
        ws.cell(row, 2, total_folders)
        row += 1
        
        ws.cell(row, 1, "Total Size:")
        ws.cell(row, 2, format_size(total_size))
        row += 2
        
        # Issues summary
        ws.cell(row, 1, "Migration Issues Found")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1
        
        for issue_name, count in issues.items():
            ws.cell(row, 1, issue_name)
            ws.cell(row, 2, count)
            if count > 0:
                ws.cell(row, 2).font = Font(color="FF0000", bold=True)
            row += 1
        
        row += 1
        
        # Department breakdown
        ws.cell(row, 1, "Department Classification")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row, 1, "Department")
        ws.cell(row, 2, "File Count")
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).font = Font(bold=True)
        row += 1
        
        for dept, count in sorted(dept_counts.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row, 1, dept)
            ws.cell(row, 2, count)
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        return ws
    
    def create_files_sheet(self, df, file_classifications):
        """Create detailed files sheet."""
        files_df = df[df['Type'] == 'File'].copy()
        
        # Add classification
        files_df['Department'] = files_df['Path'].map(file_classifications)
        
        # Select and order columns
        columns = [
            'Path', 'Name', 'Extension', 'SizeBytes', 'Created', 'LastModified',
            'Department', 'PathLength', 'HasUnsupportedChars', 'IsUnsafeExtension',
            'IsLargeFile', 'IsTooLongPath'
        ]
        
        files_df = files_df[columns]
        
        # Create sheet
        ws = self.workbook.create_sheet("Files")
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(files_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
        
        # Apply header style
        self.apply_header_style(ws)
        
        # Auto-fit columns
        self.auto_fit_columns(ws)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        return ws
    
    def create_folders_sheet(self, df, folder_classifications, folder_confidence):
        """Create folders sheet."""
        folders_df = df[df['Type'] == 'Folder'].copy()
        
        # Add classification and confidence
        folders_df['Department'] = folders_df['Path'].map(folder_classifications)
        folders_df['Confidence'] = folders_df['Path'].map(folder_confidence)
        folders_df['Confidence'] = folders_df['Confidence'].fillna(0).round(2)
        
        # Select and order columns
        columns = [
            'Path', 'Name', 'Department', 'Confidence', 'FileCountInFolder',
            'Created', 'LastModified', 'PathLength', 'HasUnsupportedChars',
            'IsTooManyFiles', 'IsTooLongPath', 'HasExplicitPermissions'
        ]
        
        folders_df = folders_df[columns]
        
        # Create sheet
        ws = self.workbook.create_sheet("Folders")
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(folders_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
        
        # Apply header style
        self.apply_header_style(ws)
        
        # Auto-fit columns
        self.auto_fit_columns(ws)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        return ws
    
    def create_permissions_sheet(self, permissions_df):
        """Create permissions sheet."""
        ws = self.workbook.create_sheet("Permissions")
        
        # Filter to only Allow access
        permissions_df = permissions_df[permissions_df['AccessControlType'] == 'Allow'].copy()
        
        # Select columns
        columns = ['Path', 'Account', 'AccessLevel']
        permissions_df = permissions_df[columns]
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(permissions_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
        
        # Apply header style
        self.apply_header_style(ws)
        
        # Auto-fit columns
        self.auto_fit_columns(ws)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        return ws
    
    def create_issues_sheet(self, df, file_classifications, folder_classifications):
        """Create consolidated issues sheet."""
        issues_list = []
        
        # Process each row
        for idx, row in df.iterrows():
            item_path = row['Path']
            item_type = row['Type']
            item_name = row['Name']
            
            # Get classification
            if item_type == 'File':
                dept = file_classifications.get(item_path, 'Unclassified')
            else:
                dept = folder_classifications.get(item_path, 'Unclassified')
            
            # Check for issues
            if row['IsTooLongPath']:
                issues_list.append({
                    'Issue Type': 'Long Path',
                    'Path': item_path,
                    'Name': item_name,
                    'Type': item_type,
                    'Department': dept,
                    'Details': f"Path length: {row['PathLength']} chars (>255)"
                })
            
            if row['IsLargeFile']:
                size_mb = row['SizeBytes'] / (1024 * 1024)
                issues_list.append({
                    'Issue Type': 'Large File',
                    'Path': item_path,
                    'Name': item_name,
                    'Type': item_type,
                    'Department': dept,
                    'Details': f"File size: {size_mb:.2f} MB (>50MB)"
                })
            
            if row['IsTooManyFiles']:
                issues_list.append({
                    'Issue Type': 'Too Many Files',
                    'Path': item_path,
                    'Name': item_name,
                    'Type': item_type,
                    'Department': dept,
                    'Details': f"File count: {row['FileCountInFolder']} (>20,000)"
                })
            
            if row['HasUnsupportedChars']:
                issues_list.append({
                    'Issue Type': 'Unsupported Characters',
                    'Path': item_path,
                    'Name': item_name,
                    'Type': item_type,
                    'Department': dept,
                    'Details': 'Name contains unsupported characters'
                })
            
            if row['IsUnsafeExtension']:
                issues_list.append({
                    'Issue Type': 'Unsafe Extension',
                    'Path': item_path,
                    'Name': item_name,
                    'Type': item_type,
                    'Department': dept,
                    'Details': f"Extension: {row['Extension']}"
                })
        
        # Convert to DataFrame
        if issues_list:
            issues_df = pd.DataFrame(issues_list)
        else:
            # Create empty DataFrame with columns
            issues_df = pd.DataFrame(columns=[
                'Issue Type', 'Path', 'Name', 'Type', 'Department', 'Details'
            ])
        
        # Create sheet
        ws = self.workbook.create_sheet("Issues")
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(issues_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
        
        # Apply header style
        self.apply_header_style(ws)
        
        # Auto-fit columns
        self.auto_fit_columns(ws)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        return ws
    
    def create_classification_sheet(self, df, folder_classifications, folder_confidence, 
                                    file_classifications):
        """Create department classification details sheet."""
        classification_list = []
        
        # Folders
        folders_df = df[df['Type'] == 'Folder']
        for idx, row in folders_df.iterrows():
            path = row['Path']
            classification_list.append({
                'Type': 'Folder',
                'Path': path,
                'Name': row['Name'],
                'Department': folder_classifications.get(path, 'Unclassified'),
                'Confidence Score': folder_confidence.get(path, 0.0)
            })
        
        # Files (sample - top by size)
        files_df = df[df['Type'] == 'File'].nlargest(1000, 'SizeBytes')
        for idx, row in files_df.iterrows():
            path = row['Path']
            classification_list.append({
                'Type': 'File',
                'Path': path,
                'Name': row['Name'],
                'Department': file_classifications.get(path, 'Unclassified'),
                'Confidence Score': ''
            })
        
        classification_df = pd.DataFrame(classification_list)
        
        # Create sheet
        ws = self.workbook.create_sheet("Classification")
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(classification_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
        
        # Apply header style
        self.apply_header_style(ws)
        
        # Auto-fit columns
        self.auto_fit_columns(ws)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        return ws
    
    def save(self):
        """Save the workbook."""
        self.workbook.save(self.output_file)
        print(f"\nExcel report saved to: {self.output_file}")


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description='Classify and generate Excel report for SharePoint migration analysis'
    )
    parser.add_argument('--config', required=True, help='Configuration JSON file')
    parser.add_argument('--raw-data', required=True, help='Raw scan data CSV file')
    parser.add_argument('--permissions', required=True, help='Permissions CSV file')
    parser.add_argument('--keywords', required=False, help='Optional: Custom department keywords CSV file (uses embedded keywords if not provided)')
    parser.add_argument('--output', required=True, help='Output Excel file')
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("SharePoint Migration Analysis - Classification & Reporting")
    print("=" * 70)
    
    # Check if input files exist
    if not Path(args.raw_data).exists():
        print(f"Error: Raw data file not found: {args.raw_data}")
        return 1
    
    if not Path(args.permissions).exists():
        print(f"Error: Permissions file not found: {args.permissions}")
        return 1
    
    # Load data
    print("\nLoading scan data...")
    df = pd.read_csv(args.raw_data)
    print(f"Loaded {len(df)} items")
    
    # Convert boolean columns
    bool_columns = [
        'HasUnsupportedChars', 'IsUnsafeExtension', 'IsLargeFile',
        'IsTooManyFiles', 'IsTooLongPath', 'HasExplicitPermissions'
    ]
    for col in bool_columns:
        df[col] = df[col].astype(bool)
    
    # Load permissions
    print("\nLoading permissions data...")
    permissions_df = pd.read_csv(args.permissions)
    print(f"Loaded {len(permissions_df)} permission entries")
    
    # Initialize classifier (with optional custom keywords file)
    classifier = DepartmentClassifier(args.keywords if args.keywords else None)
    
    # Perform classification
    folder_classifications, folder_confidence, file_classifications = \
        classifier.classify_hierarchy(df)
    
    # Generate Excel report
    print("\nGenerating Excel report...")
    report = ExcelReportGenerator(args.output)
    
    report.create_summary_sheet(df, folder_classifications, file_classifications)
    print("  - Summary sheet created")
    
    report.create_files_sheet(df, file_classifications)
    print("  - Files sheet created")
    
    report.create_folders_sheet(df, folder_classifications, folder_confidence)
    print("  - Folders sheet created")
    
    report.create_permissions_sheet(permissions_df)
    print("  - Permissions sheet created")
    
    report.create_issues_sheet(df, file_classifications, folder_classifications)
    print("  - Issues sheet created")
    
    report.create_classification_sheet(df, folder_classifications, folder_confidence,
                                       file_classifications)
    print("  - Classification sheet created")
    
    # Save workbook
    report.save()
    
    print("\n" + "=" * 70)
    print("Report generation complete!")
    print("=" * 70)
    
    return 0


if __name__ == '__main__':
    exit(main())


